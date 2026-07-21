"""
Builds a real, editable .xlsx workbook from a run's result payload -- the
exact same dict shape `app.py`'s job store already holds
(`{"jd": ..., "result": {...}, "details": {...}, "llm_warning": ...}`,
identical to what `/api/run/<job_id>/result` returns as JSON). No new
scoring/parsing logic here, ever, matching WEB_APP_PLAN.md Section 6 --
this module only serializes data that already exists into a different
format. A user can download this, edit it, filter it, re-sort it, or hand
it to someone who's never seen the web app at all.

Deliberately usable for both a live, just-finished job (today) and a
historical run once persistent run history exists (a later milestone) --
both cases just call `build_workbook()` with the same payload shape, so
there is exactly one export implementation, not two.
"""

from io import BytesIO
from urllib.parse import quote

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

HEADER_FONT = Font(bold=True)


def _autosize_columns(ws, rows: list[list], max_width: int = 60) -> None:
    """Rough column auto-sizing -- openpyxl doesn't do this automatically.
    Not visual design (Section 1 of REBUILD_PROMPT.md is about the
    web UI's look, not a spreadsheet's usability), just makes the
    downloaded file usable without the user manually resizing every
    column first."""
    if not rows:
        return
    widths = [0] * len(rows[0])
    for row in rows:
        for i, cell in enumerate(row):
            widths[i] = max(widths[i], len(str(cell)) if cell is not None else 0)
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = min(max(width + 2, 10), max_width)


def _write_sheet(ws, header: list[str], rows: list[list]) -> None:
    ws.append(header)
    for cell in ws[1]:
        cell.font = HEADER_FONT
    ws.freeze_panes = "A2"
    for row in rows:
        ws.append(row)
    _autosize_columns(ws, [header] + rows)


def _skill_match_summary(matches: list[dict] | None) -> str:
    if not matches:
        return ""
    labels = {"exact": "matched", "synonym": "matched (synonym)", "partial": "partial", "implicit": "inferred"}
    parts = []
    for m in matches:
        label = labels.get(m.get("match_type"), "not found")
        parts.append(f"{m.get('skill')}: {label}")
    return "; ".join(parts)


def _build_detail_link(file_name: str, link_params: dict | None) -> str | None:
    """
    Milestone 5.5: builds a real, shareable URL to this candidate's
    /candidate detail page -- works for anyone who opens it, not just the
    browser tab that ran the batch, because candidate.html (as of this
    milestone) can now fetch its data server-side via ?job_id= or
    ?run_id=&client_id= instead of only reading localStorage.

    Returns None (no link column added at all -- see build_workbook) if
    link_params wasn't supplied, or was supplied without enough information
    to build a working link. Never raises on a missing/malformed
    link_params -- a broken link parameter set should degrade to "no link
    column," not fail the whole export.
    """
    if not link_params:
        return None
    base = (link_params.get("base_url") or "").rstrip("/")
    if not base:
        return None
    file_q = quote(file_name, safe="")

    job_id = link_params.get("job_id")
    if job_id:
        return f"{base}/candidate?job_id={quote(job_id, safe='')}&file={file_q}"

    run_id = link_params.get("run_id")
    client_id = link_params.get("client_id")
    if run_id and client_id:
        return f"{base}/candidate?run_id={quote(run_id, safe='')}&client_id={quote(client_id, safe='')}&file={file_q}"

    return None


def _link_params_usable(link_params: dict | None) -> bool:
    """
    Whether link_params has enough information to build a real link for
    EVERY row (not just some) -- used to decide whether the "Details Link"
    column exists at all. A dict that's present but missing base_url, or
    missing both id shapes, would otherwise produce a header with no working
    links under it, which is worse than no column: it looks like a feature
    that's broken rather than a feature that wasn't requested. Mirrors
    exactly the conditions _build_detail_link itself checks, kept as a
    separate function so build_workbook can decide the column layout before
    it has any actual file names to test against.
    """
    if not link_params or not link_params.get("base_url"):
        return False
    return bool(link_params.get("job_id")) or bool(link_params.get("run_id") and link_params.get("client_id"))


def _candidate_row(row: dict, details: dict, with_cgpa_flag: bool, link_params: dict | None = None) -> list:
    detail = details.get(row.get("file"), {}) or {}
    base = [
        row.get("rank"),
        row.get("file"),
        row.get("name"),
        row.get("score"),
        row.get("confidence"),
        row.get("parse_quality"),
        "Yes" if row.get("llm_call_failed") else "No",
        " | ".join(row.get("reasons") or []),
        detail.get("college"),
        detail.get("degree_branch"),
        detail.get("graduation_year"),
        detail.get("cgpa_10pt"),
        detail.get("cgpa_source_format"),
        ", ".join(detail.get("skills") or []),
        _skill_match_summary(detail.get("required_skill_matches")),
        _skill_match_summary(detail.get("preferred_skill_matches")),
    ]
    if with_cgpa_flag:
        base.append("Yes" if row.get("below_cgpa_minimum") else "No")
    if _link_params_usable(link_params):
        # Milestone 5.5: added last, after with_cgpa_flag's column, so the
        # Reserve sheet's existing "Below CGPA Minimum" column position
        # never shifts regardless of whether a link was requested --
        # avoids silently breaking any spreadsheet-consumer (human or
        # script) that already relies on that column's position.
        base.append(_build_detail_link(row.get("file"), link_params))
    return base


def _apply_hyperlink_column(ws, col_index: int) -> None:
    """
    The URL was already written as a plain string value (via
    _write_sheet -> ws.append), same as every other column -- this just
    makes that same cell a real clickable Excel hyperlink afterward,
    without changing what _write_sheet/_autosize_columns already do for
    every other column. No-op on any cell whose value isn't a link
    (e.g. a row where _build_detail_link returned None because this
    specific file had no detail entry -- Needs Review candidates never
    reach this function at all, but a defensive check costs nothing).
    """
    for row in ws.iter_rows(min_row=2, min_col=col_index, max_col=col_index):
        cell = row[0]
        if cell.value:
            cell.hyperlink = cell.value
            cell.style = "Hyperlink"


def build_workbook(payload: dict, link_params: dict | None = None) -> Workbook:
    """
    payload shape (same as /api/run/<job_id>/result's JSON body):
        {"jd": str, "result": {"summary": {...}, "shortlist": [...],
         "reserve": [...], "needs_review": [...]}, "details": {...},
         "llm_warning": dict | None}

    link_params (Milestone 5.5, optional): when supplied, the Shortlist and
    Reserve sheets gain a "Details Link" column with a real, working URL to
    that candidate's /candidate detail page -- see _build_detail_link's
    docstring for the shape. Omitted entirely (not just left blank) when
    link_params is None/empty, so every existing caller (and every existing
    test) that doesn't pass it gets byte-for-byte the same sheets as before
    this milestone -- this is purely additive.
    """
    result = payload.get("result", {})
    details = payload.get("details", {}) or {}
    summary = result.get("summary", {})

    wb = Workbook()

    # --- Summary sheet ---
    summary_ws = wb.active
    summary_ws.title = "Summary"
    summary_rows = [
        ["Job Description", payload.get("jd")],
        ["Evaluated", summary.get("evaluated")],
        ["Shortlisted", summary.get("shortlisted")],
        ["Score cutoff used", summary.get("score_cutoff_used")],
        ["Parse failures", summary.get("parse_failures")],
        ["AI extraction failures", summary.get("llm_extraction_failures")],
    ]
    # Milestone 5.5: only present when the run actually reported it --
    # older stored payloads (a historical run persisted before this
    # milestone) won't have this key, and .get() already handles that
    # (None), but skip the row entirely rather than showing a blank
    # "Slots unfilled:" line for those.
    if summary.get("slots_unfilled") is not None:
        summary_rows.append(["Slots unfilled (no real skill-overlap candidate)", summary.get("slots_unfilled")])
    llm_warning = payload.get("llm_warning")
    if llm_warning:
        summary_rows.append(["Warning", llm_warning.get("message")])
    for row in summary_rows:
        summary_ws.append(row)
    for cell in summary_ws["A"]:
        cell.font = HEADER_FONT
    _autosize_columns(summary_ws, summary_rows, max_width=100)

    # Base columns every sheet shares, before either optional column is
    # added. The two optional columns must always be appended in this exact
    # order -- below_cgpa_minimum then details_link -- because
    # _candidate_row builds each row's list in that same order (with_cgpa_flag
    # appends first, link_params appends last), and the two must match
    # position-for-position for openpyxl to write the right value under the
    # right header.
    base_header = [
        "Rank", "File", "Name", "Score", "Confidence", "Parse Quality",
        "AI Extraction Failed", "Reasons",
        "College", "Degree/Branch", "Graduation Year", "CGPA", "CGPA Source",
        "Skills", "Required Skill Matches", "Preferred Skill Matches",
    ]
    has_links = _link_params_usable(link_params)
    link_column = ["Details Link"] if has_links else []
    candidate_header = base_header + link_column
    reserve_header = base_header + ["Below CGPA Minimum"] + link_column

    # --- Shortlist sheet ---
    shortlist_ws = wb.create_sheet("Shortlist")
    shortlist_rows = [_candidate_row(r, details, with_cgpa_flag=False, link_params=link_params) for r in result.get("shortlist", [])]
    _write_sheet(shortlist_ws, candidate_header, shortlist_rows)
    if has_links:
        _apply_hyperlink_column(shortlist_ws, len(candidate_header))

    # --- Reserve sheet ---
    reserve_ws = wb.create_sheet("Reserve")
    reserve_rows = [_candidate_row(r, details, with_cgpa_flag=True, link_params=link_params) for r in result.get("reserve", [])]
    _write_sheet(reserve_ws, reserve_header, reserve_rows)
    if has_links:
        _apply_hyperlink_column(reserve_ws, len(reserve_header))

    # --- Needs Review sheet ---
    review_ws = wb.create_sheet("Needs Review")
    review_rows = [
        [r.get("file"), r.get("reason"), "Yes" if r.get("llm_call_failed") else "No"]
        for r in result.get("needs_review", [])
    ]
    _write_sheet(review_ws, ["File", "Reason", "AI Extraction Failed"], review_rows)

    return wb


def workbook_to_bytes(wb: Workbook) -> bytes:
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
