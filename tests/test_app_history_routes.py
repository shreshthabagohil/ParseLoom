"""
Flask-integration tests for the persistent run history routes added in
Milestone 4: GET /api/runs, GET /api/runs/<run_id>, and
GET /api/runs/<run_id>/export.xlsx. Also covers that a completed
/api/run job actually gets persisted into history when a client_id is
sent, and that history is correctly per-client scoped end to end
through the real HTTP layer (not just at the src/history.py unit level).
"""
import io
import sys
import time
from pathlib import Path
from unittest.mock import patch

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import app as app_module  # noqa: E402
import src.history as history_module  # noqa: E402
from src.jobs import JobStore  # noqa: E402
from src.models import ParsedResume  # noqa: E402


def _pdf_file(name="resume.pdf"):
    return (io.BytesIO(b"%PDF-1.4 fake"), name)


def _wait_for_terminal(client, job_id, timeout=5.0):
    deadline = time.time() + timeout
    status = None
    while time.time() < deadline:
        status = client.get(f"/api/run/{job_id}/status").get_json()
        if status["status"] in ("done", "error"):
            return status
        time.sleep(0.01)
    raise TimeoutError("job did not reach a terminal state in time")


def _fake_run_batch(paths, max_workers=None, on_result=None):
    resumes = []
    for path in paths:
        import os

        resume = ParsedResume(file_name=os.path.basename(path), parse_status="Clean", full_name="Test Candidate")
        if on_result is not None:
            on_result(path, resume)
        resumes.append(resume)
    return resumes


def _run_upload_and_wait(client, client_id=None, jd="frontend"):
    data = {"resumes": [_pdf_file()], "jd": jd}
    if client_id is not None:
        data["client_id"] = client_id
    resp = client.post("/api/run", data=data, content_type="multipart/form-data")
    job_id = resp.get_json()["job_id"]
    _wait_for_terminal(client, job_id)
    return job_id


def test_completed_run_with_client_id_is_persisted_and_listed(tmp_path):
    db_path = tmp_path / "runs.db"
    client = app_module.app.test_client()
    valid_client_id = "aaaaaaaaaaaaaaaa"  # 16 chars, matches CLIENT_ID_PATTERN

    with patch.object(history_module, "DB_PATH", db_path), \
         patch("app.job_store", JobStore()), patch("app.run_batch", _fake_run_batch):
        job_id = _run_upload_and_wait(client, client_id=valid_client_id)

        list_resp = client.get(f"/api/runs?client_id={valid_client_id}")

    assert list_resp.status_code == 200
    runs = list_resp.get_json()["runs"]
    assert len(runs) == 1
    assert runs[0]["id"] == job_id


def test_run_without_client_id_is_not_persisted(tmp_path):
    """
    A client that never sends a client_id (e.g. an older cached frontend,
    or a client explicitly opting out) must not crash the job -- it just
    doesn't get history. _clean_client_id(None) -> None, and
    _execute_job's `if client_id:` guard skips the history.save_run call.
    """
    db_path = tmp_path / "runs.db"
    client = app_module.app.test_client()

    with patch.object(history_module, "DB_PATH", db_path), \
         patch("app.job_store", JobStore()), patch("app.run_batch", _fake_run_batch):
        job_id = _run_upload_and_wait(client, client_id=None)
        assert job_id  # job itself still completed fine

        # No client_id was ever associated, so nothing to list under any ID.
        list_resp = client.get("/api/runs?client_id=aaaaaaaaaaaaaaaa")

    assert list_resp.get_json()["runs"] == []


def test_get_run_returns_full_result_payload_for_the_owner(tmp_path):
    db_path = tmp_path / "runs.db"
    client = app_module.app.test_client()
    valid_client_id = "bbbbbbbbbbbbbbbb"

    with patch.object(history_module, "DB_PATH", db_path), \
         patch("app.job_store", JobStore()), patch("app.run_batch", _fake_run_batch):
        job_id = _run_upload_and_wait(client, client_id=valid_client_id)

        detail_resp = client.get(f"/api/runs/{job_id}?client_id={valid_client_id}")

    assert detail_resp.status_code == 200
    payload = detail_resp.get_json()
    assert payload["jd"]  # same shape as /api/run/<job_id>/result
    assert "result" in payload


def test_client_a_cannot_read_client_bs_run_via_api(tmp_path):
    """
    Security regression: the whole point of per-client scoping. Even
    though client A knows the real run_id (e.g. by observing network
    traffic on a shared machine, or simply guessing sequential IDs --
    not the case here since IDs are UUID4, but the ownership check must
    hold regardless), the API must refuse to serve client B's data.
    """
    db_path = tmp_path / "runs.db"
    client = app_module.app.test_client()
    client_b_id = "cccccccccccccccc"
    client_a_id = "dddddddddddddddd"

    with patch.object(history_module, "DB_PATH", db_path), \
         patch("app.job_store", JobStore()), patch("app.run_batch", _fake_run_batch):
        job_id = _run_upload_and_wait(client, client_id=client_b_id)

        # Client A tries to read client B's real run_id using their own client_id.
        cross_read_resp = client.get(f"/api/runs/{job_id}?client_id={client_a_id}")
        cross_list_resp = client.get(f"/api/runs?client_id={client_a_id}")
        cross_export_resp = client.get(f"/api/runs/{job_id}/export.xlsx?client_id={client_a_id}")

    assert cross_read_resp.status_code == 404
    assert cross_list_resp.get_json()["runs"] == []
    assert cross_export_resp.status_code == 404


def test_get_run_without_client_id_returns_404_not_the_data(tmp_path):
    db_path = tmp_path / "runs.db"
    client = app_module.app.test_client()
    valid_client_id = "eeeeeeeeeeeeeeee"

    with patch.object(history_module, "DB_PATH", db_path), \
         patch("app.job_store", JobStore()), patch("app.run_batch", _fake_run_batch):
        job_id = _run_upload_and_wait(client, client_id=valid_client_id)

        no_client_resp = client.get(f"/api/runs/{job_id}")

    assert no_client_resp.status_code == 404


def test_get_unknown_run_id_returns_404(tmp_path):
    db_path = tmp_path / "runs.db"
    client = app_module.app.test_client()
    with patch.object(history_module, "DB_PATH", db_path):
        resp = client.get("/api/runs/does-not-exist?client_id=ffffffffffffffff")
    assert resp.status_code == 404


def test_export_historical_run_returns_a_real_xlsx(tmp_path):
    db_path = tmp_path / "runs.db"
    client = app_module.app.test_client()
    valid_client_id = "1111111111111111"

    with patch.object(history_module, "DB_PATH", db_path), \
         patch("app.job_store", JobStore()), patch("app.run_batch", _fake_run_batch):
        job_id = _run_upload_and_wait(client, client_id=valid_client_id)

        export_resp = client.get(f"/api/runs/{job_id}/export.xlsx?client_id={valid_client_id}")

    assert export_resp.status_code == 200
    assert export_resp.mimetype == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert export_resp.data[:2] == b"PK"
    assert "attachment" in export_resp.headers.get("Content-Disposition", "")


def test_export_historical_run_details_link_uses_run_id_and_client_id(tmp_path):
    """
    Milestone 5.5: the historical export's "Details Link" column must point
    at /candidate?run_id=<run_id>&client_id=<client_id>&file=..., not the
    job_id-based shape the live-job export route uses -- confirms app.py's
    api_export_run_xlsx passes the right ids into build_workbook(), since
    src/output/xlsx_export.py itself can't tell live and historical apart on
    its own.
    """
    from openpyxl import load_workbook

    db_path = tmp_path / "runs.db"
    client = app_module.app.test_client()
    valid_client_id = "2222222222222222"

    with patch.object(history_module, "DB_PATH", db_path), \
         patch("app.job_store", JobStore()), patch("app.run_batch", _fake_run_batch):
        job_id = _run_upload_and_wait(client, client_id=valid_client_id)
        export_resp = client.get(f"/api/runs/{job_id}/export.xlsx?client_id={valid_client_id}")

    wb = load_workbook(io.BytesIO(export_resp.data))
    sheet = wb["Shortlist"] if wb["Shortlist"].max_row > 1 else wb["Reserve"]
    header = [c.value for c in sheet[1]]
    assert "Details Link" in header
    link_col = header.index("Details Link") + 1
    link_value = sheet.cell(row=2, column=link_col).value
    assert link_value is not None
    assert f"run_id={job_id}" in link_value
    assert f"client_id={valid_client_id}" in link_value
    assert "job_id=" not in link_value


def test_malformed_client_id_is_rejected_not_silently_accepted(tmp_path):
    """
    _clean_client_id() enforces CLIENT_ID_PATTERN (alnum/-/_ , 8-64
    chars). A client_id containing SQL-metacharacter-adjacent or
    otherwise out-of-pattern content should be treated the same as "no
    client_id" -- an empty list / 404, never passed through to a query.
    """
    db_path = tmp_path / "runs.db"
    client = app_module.app.test_client()
    with patch.object(history_module, "DB_PATH", db_path):
        list_resp = client.get("/api/runs?client_id=" + "a" * 3)  # too short
        detail_resp = client.get("/api/runs/some-run?client_id=has spaces!!")

    assert list_resp.get_json()["runs"] == []
    assert detail_resp.status_code == 404


def test_list_runs_with_no_client_id_returns_empty_list_not_error(tmp_path):
    db_path = tmp_path / "runs.db"
    client = app_module.app.test_client()
    with patch.object(history_module, "DB_PATH", db_path):
        resp = client.get("/api/runs")
    assert resp.status_code == 200
    assert resp.get_json()["runs"] == []
