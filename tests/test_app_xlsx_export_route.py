"""
Flask-integration tests for GET /api/run/<job_id>/export.xlsx.
"""
import io
import sys
import time
from pathlib import Path
from unittest.mock import patch

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import app as app_module  # noqa: E402
from src.jobs import JobStore  # noqa: E402
from src.models import ParsedResume  # noqa: E402


def _pdf_file(name="resume.pdf"):
    return (io.BytesIO(b"%PDF-1.4 fake"), name)


def _wait_for_terminal(client, job_id, timeout=5.0):
    deadline = time.time() + timeout
    status = None
    while time.time() < deadline:
        status = client.get(f"/api/run/{job_id}/status").get_json()
        if status["status"] in ("done", "error"):
            return status
        time.sleep(0.01)
    raise TimeoutError("job did not reach a terminal state in time")


def _fake_run_batch(paths, max_workers=None, on_result=None):
    resumes = []
    for path in paths:
        import os

        resume = ParsedResume(file_name=os.path.basename(path), parse_status="Clean", full_name="Test Candidate")
        if on_result is not None:
            on_result(path, resume)
        resumes.append(resume)
    return resumes


def test_export_returns_a_real_xlsx_for_a_completed_job():
    client = app_module.app.test_client()
    with patch("app.job_store", JobStore()), patch("app.run_batch", _fake_run_batch):
        resp = client.post(
            "/api/run", data={"resumes": [_pdf_file()], "jd": "frontend"}, content_type="multipart/form-data"
        )
        job_id = resp.get_json()["job_id"]
        _wait_for_terminal(client, job_id)

        export_resp = client.get(f"/api/run/{job_id}/export.xlsx")

    assert export_resp.status_code == 200
    assert export_resp.mimetype == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert export_resp.data[:2] == b"PK"  # real xlsx (zip) magic bytes
    assert "attachment" in export_resp.headers.get("Content-Disposition", "")


def test_export_shortlist_details_link_points_at_this_job_id():
    """
    Milestone 5.5: the exported workbook's "Details Link" column should
    point at /candidate?job_id=<this job's id>&file=<candidate file> --
    confirms app.py actually wires request.host_url + job_id into
    build_workbook(), not just that build_workbook() itself can do it
    (already covered in isolation by tests/test_xlsx_export.py).
    """
    from openpyxl import load_workbook

    client = app_module.app.test_client()
    with patch("app.job_store", JobStore()), patch("app.run_batch", _fake_run_batch):
        resp = client.post(
            "/api/run", data={"resumes": [_pdf_file("candidate_x.pdf")], "jd": "frontend"}, content_type="multipart/form-data"
        )
        job_id = resp.get_json()["job_id"]
        _wait_for_terminal(client, job_id)
        export_resp = client.get(f"/api/run/{job_id}/export.xlsx")

    wb = load_workbook(io.BytesIO(export_resp.data))
    sheet = wb["Shortlist"] if wb["Shortlist"].max_row > 1 else wb["Reserve"]
    header = [c.value for c in sheet[1]]
    assert "Details Link" in header
    link_col = header.index("Details Link") + 1
    link_value = sheet.cell(row=2, column=link_col).value
    assert link_value is not None
    assert f"job_id={job_id}" in link_value
    assert "file=candidate_x.pdf" in link_value


def test_export_unknown_job_returns_404():
    client = app_module.app.test_client()
    with patch("app.job_store", JobStore()):
        resp = client.get("/api/run/does-not-exist/export.xlsx")
    assert resp.status_code == 404


def test_export_still_running_returns_409():
    def _slow_run_batch(paths, max_workers=None, on_result=None):
        time.sleep(0.3)
        return _fake_run_batch(paths, max_workers, on_result)

    client = app_module.app.test_client()
    store = JobStore()
    with patch("app.job_store", store), patch("app.run_batch", _slow_run_batch):
        resp = client.post(
            "/api/run", data={"resumes": [_pdf_file()], "jd": "frontend"}, content_type="multipart/form-data"
        )
        job_id = resp.get_json()["job_id"]
        export_resp = client.get(f"/api/run/{job_id}/export.xlsx")
        _wait_for_terminal(client, job_id)  # drain before patch context exits

    assert export_resp.status_code == 409


def test_export_filename_is_sanitized_against_header_injection():
    """
    Security regression: the download filename is built from the JD
    "role" string, which can originate from an LLM-extracted free-text JD
    (attacker-influenced content, same trust boundary as resume text).
    Must never allow characters that could manipulate the
    Content-Disposition header or write outside a safe filename.
    """
    import src.jd_loader as jd_loader_module

    malicious_role_jd = {
        "role": '../../evil"; filename="hacked.txt',
        "required_skills": ["Python"],
        "preferred_skills": [],
        "cgpa_min": 0,
        "slots": 1,
    }

    client = app_module.app.test_client()
    with patch("app.job_store", JobStore()), \
         patch("app.run_batch", _fake_run_batch), \
         patch.object(jd_loader_module, "load", return_value=jd_loader_module.JobDescription(**malicious_role_jd)):
        resp = client.post(
            "/api/run", data={"resumes": [_pdf_file()], "jd": "anything"}, content_type="multipart/form-data"
        )
        job_id = resp.get_json()["job_id"]
        _wait_for_terminal(client, job_id)
        export_resp = client.get(f"/api/run/{job_id}/export.xlsx")

    assert export_resp.status_code == 200
    disposition = export_resp.headers.get("Content-Disposition", "")
    assert ".." not in disposition
    assert "/" not in disposition
    # No stray, unescaped quote that could break out of the filename="..."
    # attribute and inject additional header content.
    assert disposition.count('"') <= 2
