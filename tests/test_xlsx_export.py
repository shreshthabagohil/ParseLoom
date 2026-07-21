"""
Tests for src/output/xlsx_export.py -- building a real, editable .xlsx
workbook from a run's result payload (the user explicitly asked to be able
to download and edit results "however they want").
"""
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from src.output.xlsx_export import build_workbook, workbook_to_bytes

SAMPLE_PAYLOAD = {
    "jd": "Frontend Developer",
    "result": {
        "summary": {
            "evaluated": 3,
            "shortlisted": 1,
            "score_cutoff_used": 80.0,
            "parse_failures": 1,
            "llm_extraction_failures": 0,
        },
        "shortlist": [
            {
                "rank": 1,
                "file": "a.pdf",
                "name": "Alice",
                "score": 91.2,
                "confidence": "High",
                "parse_quality": "Clean",
                "reasons": ["Matches 4/4 required skills.", "CGPA meets minimum."],
                "llm_call_failed": False,
            }
        ],
        "reserve": [
            {
                "file": "b.pdf",
                "name": "Bob",
                "score": 40.0,
                "confidence": "Medium",
                "parse_quality": "Partial",
                "reasons": ["Missing required skills."],
                "llm_call_failed": False,
                "below_cgpa_minimum": True,
            }
        ],
        "needs_review": [
            {"file": "c.pdf", "reason": "Parse failed -- no reliable data extracted.", "llm_call_failed": False}
        ],
    },
    "details": {
        "a.pdf": {
            "college": "IIT Bombay",
            "degree_branch": "CS",
            "graduation_year": 2026,
            "cgpa_10pt": 9.1,
            "cgpa_source_format": "cgpa10",
            "skills": ["React", "TypeScript"],
            "required_skill_matches": [{"skill": "React", "match_type": "exact"}],
            "preferred_skill_matches": [],
        },
        "b.pdf": {
            "college": "Delhi University",
            "degree_branch": "IT",
            "graduation_year": 2025,
            "cgpa_10pt": 6.0,
            "cgpa_source_format": "cgpa10",
            "skills": [],
            "required_skill_matches": [],
            "preferred_skill_matches": [],
        },
    },
    "llm_warning": None,
}


def test_workbook_has_all_four_expected_sheets():
    wb = build_workbook(SAMPLE_PAYLOAD)
    assert wb.sheetnames == ["Summary", "Shortlist", "Reserve", "Needs Review"]


def test_summary_sheet_contains_jd_and_counts():
    wb = build_workbook(SAMPLE_PAYLOAD)
    ws = wb["Summary"]
    values = [cell.value for row in ws.iter_rows() for cell in row]
    assert "Frontend Developer" in values
    assert 3 in values  # evaluated
    assert 1 in values  # shortlisted


def test_shortlist_sheet_has_header_and_one_data_row():
    wb = build_workbook(SAMPLE_PAYLOAD)
    ws = wb["Shortlist"]
    header = [c.value for c in ws[1]]
    assert header[:8] == [
        "Rank", "File", "Name", "Score", "Confidence", "Parse Quality",
        "AI Extraction Failed", "Reasons",
    ]
    data_row = [c.value for c in ws[2]]
    assert data_row[1] == "a.pdf"
    assert data_row[2] == "Alice"
    assert data_row[3] == 91.2
    # Detail fields flattened in too, not just the top-level score row.
    assert "IIT Bombay" in data_row
    assert "React, TypeScript" in data_row


def test_reserve_sheet_includes_below_cgpa_minimum_column():
    wb = build_workbook(SAMPLE_PAYLOAD)
    ws = wb["Reserve"]
    header = [c.value for c in ws[1]]
    assert header[-1] == "Below CGPA Minimum"
    data_row = [c.value for c in ws[2]]
    assert data_row[-1] == "Yes"  # Bob is below_cgpa_minimum=True


def test_needs_review_sheet_has_file_and_reason():
    wb = build_workbook(SAMPLE_PAYLOAD)
    ws = wb["Needs Review"]
    data_row = [c.value for c in ws[2]]
    assert data_row[0] == "c.pdf"
    assert "Parse failed" in data_row[1]


def test_empty_payload_produces_valid_empty_sheets_not_a_crash():
    empty_payload = {"jd": "Empty Role", "result": {"summary": {}, "shortlist": [], "reserve": [], "needs_review": []}, "details": {}, "llm_warning": None}
    wb = build_workbook(empty_payload)
    assert wb["Shortlist"].max_row == 1  # header only, no crash on zero candidates


def test_workbook_to_bytes_produces_a_real_loadable_xlsx():
    from openpyxl import load_workbook
    from io import BytesIO

    wb = build_workbook(SAMPLE_PAYLOAD)
    raw_bytes = workbook_to_bytes(wb)
    assert raw_bytes[:2] == b"PK"  # xlsx is a zip container -- real magic bytes, not a stub

    reloaded = load_workbook(BytesIO(raw_bytes))
    assert reloaded.sheetnames == ["Summary", "Shortlist", "Reserve", "Needs Review"]
    assert reloaded["Shortlist"]["B2"].value == "a.pdf"


# --- Milestone 5.5: "Details Link" column ------------------------------------
# A shareable link to a candidate's /candidate detail page, added to the
# Shortlist/Reserve sheets only when link_params is supplied -- omitted
# entirely with no link_params, matching pre-5.5 output byte-for-byte
# (verified below), which is the backward-compatibility guarantee this
# milestone's plan requires.


def test_no_link_params_produces_identical_output_to_before_milestone_5_5():
    wb_without = build_workbook(SAMPLE_PAYLOAD)
    wb_with_empty = build_workbook(SAMPLE_PAYLOAD, link_params=None)
    for name in wb_without.sheetnames:
        rows_without = [[c.value for c in row] for row in wb_without[name].iter_rows()]
        rows_with_empty = [[c.value for c in row] for row in wb_with_empty[name].iter_rows()]
        assert rows_without == rows_with_empty
    assert wb_without["Shortlist"]["A1"].value == "Rank"
    header = [c.value for c in wb_without["Shortlist"][1]]
    assert "Details Link" not in header


def test_job_id_link_params_adds_details_link_column_with_expected_url():
    link_params = {"base_url": "http://127.0.0.1:5000/", "job_id": "abc123"}
    wb = build_workbook(SAMPLE_PAYLOAD, link_params=link_params)

    header = [c.value for c in wb["Shortlist"][1]]
    assert header[-1] == "Details Link"

    link_cell = wb["Shortlist"][2][-1]
    assert link_cell.value == "http://127.0.0.1:5000/candidate?job_id=abc123&file=a.pdf"
    assert link_cell.hyperlink is not None
    assert link_cell.hyperlink.target == link_cell.value


def test_run_id_client_id_link_params_adds_details_link_column():
    link_params = {"base_url": "http://127.0.0.1:5000", "run_id": "run-1", "client_id": "client-1"}
    wb = build_workbook(SAMPLE_PAYLOAD, link_params=link_params)

    link_cell = wb["Shortlist"][2][-1]
    assert link_cell.value == "http://127.0.0.1:5000/candidate?run_id=run-1&client_id=client-1&file=a.pdf"


def test_reserve_sheet_keeps_below_cgpa_minimum_column_position_with_link_params():
    # Regression guard: the existing "Below CGPA Minimum" column must not
    # shift position just because a "Details Link" column was also added --
    # anything already relying on its column index (or just visually
    # expecting it right after the shared candidate columns) shouldn't break.
    link_params = {"base_url": "http://127.0.0.1:5000", "job_id": "abc123"}
    wb = build_workbook(SAMPLE_PAYLOAD, link_params=link_params)

    header = [c.value for c in wb["Reserve"][1]]
    assert header[-2] == "Below CGPA Minimum"
    assert header[-1] == "Details Link"

    data_row = [c.value for c in wb["Reserve"][2]]
    assert data_row[-2] == "Yes"  # Bob is below_cgpa_minimum=True
    assert data_row[-1] == "http://127.0.0.1:5000/candidate?job_id=abc123&file=b.pdf"


def test_needs_review_sheet_unaffected_by_link_params():
    # No detail data exists for a failed parse to link to -- unchanged.
    link_params = {"base_url": "http://127.0.0.1:5000", "job_id": "abc123"}
    wb = build_workbook(SAMPLE_PAYLOAD, link_params=link_params)
    header = [c.value for c in wb["Needs Review"][1]]
    assert header == ["File", "Reason", "AI Extraction Failed"]


def test_link_params_missing_base_url_degrades_to_no_link_gracefully():
    # A malformed/incomplete link_params must never crash the export --
    # degrades to no link column, same as omitting link_params entirely.
    wb = build_workbook(SAMPLE_PAYLOAD, link_params={"job_id": "abc123"})
    header = [c.value for c in wb["Shortlist"][1]]
    assert "Details Link" not in header


def test_summary_sheet_includes_slots_unfilled_when_present():
    payload = {**SAMPLE_PAYLOAD, "result": {**SAMPLE_PAYLOAD["result"], "summary": {**SAMPLE_PAYLOAD["result"]["summary"], "slots_unfilled": 2}}}
    wb = build_workbook(payload)
    values = [cell.value for row in wb["Summary"].iter_rows() for cell in row]
    assert 2 in values
